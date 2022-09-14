# coding=utf-8
from pathlib import Path
from openpyxl import Workbook,load_workbook
from tqdm import tqdm 
import pandas as pd
import sys
import numpy as np
from PIL import Image

name_str="谢家冬"
date_str="填表时间：2022年9月  日"

doc='''
\n本文件为家冬专用Python源程序。在powershell等命令行提示符后输入python  allinone.py  参数
说明：参数为\n
1-->功能为建立目录和copy文件。\n
2-->提取汇总表。 \n
3-->识别.\images文件夹的图片中文字\n
Power by qllwx@live.com 2022.09.12     
'''
def read_zdzl(filename, range="B16"):
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    return ws[range].value

def mk_first_dir(filename,data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    zdzl=read_zdzl(filename)
    dir_1=Path(data_dir,base_name)    
    mk_dir(dir_1)
    dir_2=Path(data_dir,base_name,zdzl)
    mk_dir(dir_2) 
    for i in range(1,4):
        dir_3=Path(dir_2,"问题"+str(i))
        mk_dir(dir_3)
    return dir_1

def copy_range(worksheet,source_range="B11:B25",target_range="C11:C25"):
    s=[]
    for cell in worksheet[source_range]:
        v=list(cell)[0].value
        s.append(v)
        
    idx=0
    for cell in worksheet[target_range]:        
        cell[0].value=s[idx]
        idx+=1
    return worksheet

def write_end(filename,name_str=name_str,name_range="B29",date_str=date_str,date_range="C29",data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    write_filename=Path(data_dir,base_name,Path(filename).name)
    if write_filename.exists():
        print("文件：%s-->已经存在，要重写请先删除"%write_filename)
        return
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    ws[name_range]=name_str
    ws[date_range]=date_str
    ws=copy_range(ws)

    
    print("写入：",write_filename,name_str,date_str)
    if Path(write_filename).exists():
        Path(write_filename).unlink()
    wb.save(write_filename)

def copy_curfile2_datadir(data_dir='./data/'):
    print("将当前目录下的.xlsx文件copy到指定文件夹%s中"%data_dir)
    file_list=[f for f in Path('.').iterdir() if Path(f).is_file()]
    for file in tqdm(file_list,colour='pink'):
        if (file.name.endswith('.xlsx') and not Path(file).exists()):
            src=file
            dest=Path(data_dir,file.name)
            dest.write_bytes(src.read_bytes()) #for binary files

def mk_dir(path):
    if not Path(path).exists():
        print("建立%s..."%path)
        Path.mkdir(Path(path))
        print(path,"OK!") 
    else:
        print("目录:%s-->已经存在。"%path)

def read_table(filename):
    d={}
    wb=load_workbook(filename)
    ws=wb['Sheet1']    
    d['宗地代码（不动产单元号）']=ws['C11'].value
    zdzl=str(ws['C16'].value)
    county,towns,village,LandParcelNumber =parse_zdzl(zdzl)
    d['乡（镇）']=towns
    d['村']=village
    d['宗地号']=LandParcelNumber
    d['存在问题']=ws['B26'].value
    d['现场核实情况']=ws['C26'].value
    d['处理意见']=ws['B27'].value
    return d
    
def parse_zdzl(s):
    county=s[:s.index('县')]
    if '乡' in s:
        towns=s[s.index('县')+1:s.index('乡')]
        village=s[s.index('乡')+1:s.index('村')]
    else:
        towns=s[s.index('县')+1:s.index('镇')]
        village=s[s.index('镇')+1:s.index('村')]
    LandParcelNumber=s[s.index('村')+1:s.index('号')]
    return (county,towns,village,LandParcelNumber)

def process_1(data_dir='./data/'):    
    mk_dir(data_dir) 
    copy_curfile2_datadir()
    file_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_file() and f.name.endswith('.xlsx') and len(f.name)==24)]
    print("共找到%s个数据文件"%len(file_list))
    for file in tqdm(file_list,colour='red',):
        print("读取文件",file)  
        mk_first_dir(Path(file))          
        write_end(file)

def process_2(data_dir='./data/'):
    result=None
    dir_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_dir() and  len(f.name)==19)]  
    pbar = tqdm(dir_list)
    for dir in pbar:
        pbar.set_description("\n%s"%(dir.name))
        print("%s"%dir)
        file=Path(dir,dir.name+".xlsx")
        series=pd.Series(read_table(file))
        print(series)
        if result is None:
            result = pd.DataFrame(series)
        else:
            result=pd.concat([result,series],ignore_index=True,axis=1)
    result=result.T
    result=result.set_index(result.columns[0])
    savefiles_count=len(sorted(Path('.').glob("汇总表*.xlsx")))
    savefile_name="汇总表%s.xlsx"%savefiles_count
    result.to_excel(savefile_name,index=False)
    print(result)


def get_1_4_image(img_file):   
    img=Image.open(img_file)
    w,h=img.size
    r_w=int(w*2/3)
    r_h=int(h*3/4)
    box=(0,r_h,r_w,h)
    im_crop= img.crop(box)
    im_crop=red20_other2b(im_crop)
    dest='temp.png'
    res_img = Image.fromarray(im_crop)
    res_img.save(dest)    
    #res_img.show()
    return res_img

def red20_other2b(img):	
    ''' 通过遍历颜色替换程序,只留红色，其余为白
	@param	img:	图像矩阵
	@return				替换后的图像矩阵
	'''
    img_arr = np.asarray(img, dtype=np.double)    
    for i in range(img_arr.shape[1]):
        for j in range(img_arr.shape[0]):
            r,g,b=img_arr[j][i]
            if r>(g+b)*1.8 :
               img_arr[j][i] = (255,0,0)
            else:
                img_arr[j][i]=(0,0,0)
    return np.asarray(img_arr, dtype=np.uint8)
			

def get_description(img=None,list_dir='.\images'):
    if img is None:
        dir_list=[f for f in Path(list_dir).glob("**/*") if (f.is_file() and (f.suffix in ['.jpg','.png']))] 
        result=[]
        for file in tqdm(dir_list):
            src=file
            #dest=Path('.','temp'+file.suffix)
            #dest.write_bytes(src.read_bytes()) #for binary files
            get_1_4_image(file)
            print("识别%s"%str(file))
            try:
                #res=reader.readtext(str(dest),detail=0)
                res = reader.readtext('temp.png',detail=0)
            except :
                print("ERROR")
                continue
            res=get_jwd(res)
            print(res)
            res.append(str(file))
            if len(res)==4:
                result.append(res)
    else:
        im =get_1_4_image(img)
        res = reader.readtext("temp.png",detail=0)
        res=get_jwd(res)
        res.append(img)
        print(res)
        result=[res]

    return result

def arr2dict(r):
    tmp={}
    tmp['宗地代码（不动产单元号）']=str(Path(r[-1]).parent).split('\\')[1]
    tmp['经度']=r[0]
    tmp['纬度']=r[1]
    tmp['问题']=r[2]
    return tmp


def result2df(res):    
    df=pd.DataFrame({},columns=['宗地代码（不动产单元号）','经度','纬度','问题'])
    tmp={}
    for r in res:
       tmp=arr2dict(r)
       df1=pd.DataFrame(pd.Series(tmp)).T
       df=pd.concat([df,df1],axis=0)
    #df.index=['宗地代码（不动产单元号）']
    return df

def hzb_add_by_index(df_add):
    excel_file="汇总_add_经纬度.xlsx"
    savefiles_count=len(sorted(Path('.').glob("汇总表*.xlsx")))
    savefile_name="汇总表%s.xlsx"%(savefiles_count-1)
    hzb=pd.read_excel(savefile_name)
    hzb.set_index(['宗地代码（不动产单元号）'],inplace=True)    
    df=pd.merge(hzb,df_add,on=['宗地代码（不动产单元号）'],how='outer') # left , right , outer
    df.to_excel(excel_file,index=False)
    add_columns_by_question(excel_file)
    return df

def add_columns_by_question(excel_file):
    df=pd.read_excel(excel_file)
    df1=df[df['问题'].str.contains("问题1",na=False)]
    df1.set_index('宗地代码（不动产单元号）',inplace=True)
    df2=df[df['问题'].str.contains("问题2",na=False)]
    df2=df2[['宗地代码（不动产单元号）','经度','纬度','问题']]    
    df2.columns=['宗地代码（不动产单元号）','经度2','纬度2','问题2']
    df2.set_index('宗地代码（不动产单元号）',inplace=True)
    df3=df[df['问题'].str.contains("问题3",na=False)]
    df3=df3[['宗地代码（不动产单元号）','经度','纬度','问题']]    
    df3.columns=['宗地代码（不动产单元号）','经度3','纬度3','问题3']
    df3.set_index('宗地代码（不动产单元号）',inplace=True)
    res=pd.merge(df1,df2,on=['宗地代码（不动产单元号）'],how='outer')
    res=pd.merge(res,df3,on=['宗地代码（不动产单元号）'],how='outer')
    res.to_excel("汇总_add_经纬度2.xlsx",index=False)
    return res




def get_jwd(res):
    regx_str="经纬度：；; :备注虞~^筇臣与_赓'`轻瘘。孥W%@"   
    result=[] 
    for r in res:
        for c in regx_str:
            r=r.replace(c,"")
        if len(r)>0:
            result.append(r)
    return result

def put_images(img_dir='images'):
    dir_list=[f for f in Path(img_dir).glob("**/*") if (f.is_file() and (f.suffix in ['.jpg','.png']))] 
   
    result=[]
    for img in dir_list:
        res=get_description(img=img)[0]
        result.append(res)
    df=pd.DataFrame(res)
    df.to_excel(str(Path(img_dir,'list.xlsx')),index=False)
    




if __name__=="__main__": 
    if len(sys.argv)==2:
        num=str(sys.argv[1])
    else:
        prog_name=Path(sys.argv[0]).name 
        num=prog_name[0]

    if num=='1': 
       process_1() 
    elif num=='2': 
       process_2()
    elif num=='3': 
        
        import easyocr 
        reader = easyocr.Reader(['ch_sim','en'],recog_network ='zh_sim_g2') # this needs to run only once to load the model into memory
        result =get_description(list_dir='.\data')
        res_df =result2df(result)
        print(res_df)
        hzb_add_by_index(res_df)
    elif num=="putimg":
        import easyocr 
        reader = easyocr.Reader(['ch_sim','en'],recog_network ='zh_sim_g2') # this needs to run only once to load the model into memory
        put_images()


        
    else:
        print("Usage: \n%s"%doc)