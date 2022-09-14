doc='''
修改当前目录下所有xlsx文件中的单元格内容
Usage:
python edit_xlsx.py 单元格[B26]   [原来的内容]  新的内容

Example:
python edit_xlsx.py B29 谢家冬  谢家冬、马倩雯
'''
from pathlib import Path
from openpyxl import Workbook,load_workbook
import sys
from tqdm import tqdm

def get_xlsx():
    result=[f for f in Path('.').glob('**/*.xlsx') ]
    return result

def edit_xlsx(xls_file,range,s,o):

    wb=load_workbook(xls_file)
    ws=wb['Sheet1']
    try:
        ws_str=ws[range].value
    except:
        print("第一个参数->单元格名称:%s无效"%range)
        exit()
    if (ws_str==s ) or s=="*" :        
        ws[range].value=o
        wb.save(xls_file)
        print("%s已经将'%s'处的内容'%s'修改为'%s'"%(xls_file,range,ws_str,o))
    else:
        print("%s\n源内容'%s'与指定的要修改内容'%s'不一致，未能操作"%(xls_file,ws_str,s))
   


def edit_xlsxs(files,range="B26",s=None,o=None):
    
    for f in tqdm(files):
        edit_xlsx(f,range,s,o)

def main(argv):
    if len(argv)<3:
        print(doc)
    else:
        range=argv[1]
        s=argv[2]
        o=argv[-1]
        if len(argv)==3:
            s=''
        files=get_xlsx()
        print('准备修改当前目录下的所有.xlsx文件中的"%s"内容"%s"为"%s"'%(range,s,o))
        edit_xlsxs(files,range=range,s=s,o=o)

    
    
if __name__=="__main__":
    main(sys.argv)
