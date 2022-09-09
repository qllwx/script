from pathlib import Path
from openpyxl import Workbook,load_workbook



def read_zdzl(filename, range="B16"):
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    return ws[range].value

def mk_dir(filename):
    base_name=Path(filename).name.split(".")[0]
    zdzl=read_zdzl(filename)
    dir_1=Path(data_dir,base_name)
    if not dir_1.exists():
        Path.mkdir(dir_1)
        print(dir_1,"OK!")
    dir_2=Path(data_dir,base_name,zdzl)
    if not dir_2.exists():
        Path.mkdir(dir_2)
        print(dir_2,"OK!")    
    for i in range(1,4):
        dir_3=Path(dir_2,"问题"+str(i))
        if not dir_3.exists():
           Path.mkdir(dir_3)
           print(dir_3,"OK!")
    return dir_1

def copy_range(worksheet,source_range="B11:B25",target_range="C11:C25"):
    s=[]
    for cell in worksheet[source_range]:
        v=list(cell)[0].value
        s.append(v)
        
    idx=0
    for cell in worksheet[target_range]:        
        cell[0].value=s[idx]
        idx+=1
    return worksheet

def write_end(filename,name_str="谢家冬",name_range="B29",date_str="填表时间：2022年9月  日",date_range="C29"):
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    ws[name_range]=name_str
    ws[date_range]=date_str
    ws=copy_range(ws)

    base_name=Path(filename).name.split(".")[0]
    write_filename=Path(data_dir,base_name,Path(filename).name)
    print("写入：",write_filename,name_str,date_str)
    if Path(write_filename).exists():
        Path(write_filename).unlink()
    wb.save(write_filename)

def copy_curfile2_datadir(data_dir='./data/'):
    file_list=[f for f in Path('.').iterdir() if Path(f).is_file()]
    for file in file_list:
        if file.name.endswith('.xlsx'):
            src=file
            dest=Path(data_dir,file.name)
            dest.write_bytes(src.read_bytes()) #for binary files
    





if __name__=="__main__":   
    
    data_dir='./data/'
    if not Path(data_dir).exists():
        Path.mkdir(Path(data_dir))
        print(data_dir,"OK!")  
    copy_curfile2_datadir()
    file_list=[f for f in Path(data_dir).iterdir() if Path(f).is_file()]
    for file in file_list:
        print("读取文件",file)    
        print('建立%s目录及其子目录已完成'%(mk_dir(file)))
        write_end(file)