from pathlib import Path
from openpyxl import Workbook,load_workbook
from tqdm import tqdm 
import pandas as pd
import sys

doc='''
\n本文件为家冬专用Python源程序。在powershell等命令行提示符后输入python 
说明：程序文件名中的第1个字符为\n
1-->功能为建立目录和copy文件。\n
2-->提取汇总表。 \n\n
Power by qllwx@live.com 2022.09.09     
'''
def read_zdzl(filename, range="B16"):
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    return ws[range].value

def mk_first_dir(filename,data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    zdzl=read_zdzl(filename)
    dir_1=Path(data_dir,base_name)    
    mk_dir(dir_1)
    dir_2=Path(data_dir,base_name,zdzl)
    mk_dir(dir_2) 
    for i in range(1,4):
        dir_3=Path(dir_2,"问题"+str(i))
        mk_dir(dir_3)
    return dir_1

def copy_range(worksheet,source_range="B11:B25",target_range="C11:C25"):
    s=[]
    for cell in worksheet[source_range]:
        v=list(cell)[0].value
        s.append(v)
        
    idx=0
    for cell in worksheet[target_range]:        
        cell[0].value=s[idx]
        idx+=1
    return worksheet

def write_end(filename,name_str="谢家冬",name_range="B29",date_str="填表时间：2022年9月  日",date_range="C29",data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    write_filename=Path(data_dir,base_name,Path(filename).name)
    if write_filename.exists():
        print("文件：%s-->已经存在，要重写请先删除"%write_filename)
        return
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    ws[name_range]=name_str
    ws[date_range]=date_str
    ws=copy_range(ws)

    
    print("写入：",write_filename,name_str,date_str)
    if Path(write_filename).exists():
        Path(write_filename).unlink()
    wb.save(write_filename)

def copy_curfile2_datadir(data_dir='./data/'):
    print("将当前目录下的.xlsx文件copy到指定文件夹%s中"%data_dir)
    file_list=[f for f in Path('.').iterdir() if Path(f).is_file()]
    for file in tqdm(file_list,colour='pink'):
        if (file.name.endswith('.xlsx') and not Path(file).exists()):
            src=file
            dest=Path(data_dir,file.name)
            dest.write_bytes(src.read_bytes()) #for binary files

def mk_dir(path):
    if not Path(path).exists():
        print("建立%s..."%path)
        Path.mkdir(Path(path))
        print(path,"OK!") 
    else:
        print("目录:%s-->已经存在。"%path)

def read_table(filename):
    d={}
    wb=load_workbook(filename)
    ws=wb['Sheet1']    
    d['宗地代码（不动产单元号）']=ws['C11'].value
    zdzl=str(ws['C16'].value)
    county,towns,village,LandParcelNumber =parse_zdzl(zdzl)
    d['乡（镇）']=towns
    d['村']=village
    d['宗地号']=LandParcelNumber
    d['存在问题']=ws['B26'].value
    d['现场核实情况']=ws['C26'].value
    d['处理意见']=ws['B27'].value
    return d
    
def parse_zdzl(s):
    county=s[:s.index('县')]
    if '乡' in s:
        towns=s[s.index('县')+1:s.index('乡')]
        village=s[s.index('乡')+1:s.index('村')]
    else:
        towns=s[s.index('县')+1:s.index('镇')]
        village=s[s.index('镇')+1:s.index('村')]
    LandParcelNumber=s[s.index('村')+1:s.index('号')]
    return (county,towns,village,LandParcelNumber)

def process_1(data_dir='./data/'):    
    mk_dir(data_dir) 
    copy_curfile2_datadir()
    file_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_file() and f.name.endswith('.xlsx') and len(f.name)==24)]
    print("共找到%s个数据文件"%len(file_list))
    for file in tqdm(file_list,colour='red',):
        print("读取文件",file)  
        mk_first_dir(Path(file))          
        write_end(file)

def process_2(data_dir='./data/'):
    result=None
    dir_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_dir() and  len(f.name)==19)]  
    pbar = tqdm(dir_list)
    for dir in pbar:
        pbar.set_description("\n%s"%(dir.name))
        print("%s"%dir)
        file=Path(dir,dir.name+".xlsx")
        series=pd.Series(read_table(file))
        print(series)
        if result is None:
            result = pd.DataFrame(series)
        else:
            result=pd.concat([result,series],ignore_index=True,axis=1)
    result=result.T
    result=result.set_index(result.columns[0])
    savefiles_count=len(sorted(Path('.').glob("汇总表*.xlsx")))
    savefile_name="汇总表%s.xlsx"%savefiles_count
    result.to_excel(savefile_name)
    print(result)







if __name__=="__main__": 
    prog_name=Path(sys.argv[0]).name 
    num=prog_name[0]
    if num=='1': 
       process_1() 
    elif num=='2': 
       process_2()
    else:
        print("Usage: \n%s"%doc)