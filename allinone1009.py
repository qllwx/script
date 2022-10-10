# coding=utf-8
from pathlib import Path
import re
from openpyxl import Workbook,load_workbook
from tqdm import tqdm 
import pandas as pd
import sys,shutil
import numpy as np
from PIL import Image
import re

name_str="张永红、马倩雯"
date_str="填表时间：2022.9."

doc='''
\n本文件为家冬专用Python源程序。在powershell等命令行提示符后输入python  allinone.py  参数
说明：参数为\n
1-->功能为建立目录和copy文件。含表内copy及填写 填表人和填表时间\n
2-->提取汇总表。 \n
3-->识别.\data文件夹的图片中文字+生成经纬度汇总表\n
4-->识别.\images文件夹的图片中文字，并移动到应文件夹，同时生成结果表\n

Power by qllwx@live.com 2022.10.09    
'''
def read_zdzl(filename, range="B16"):
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    return ws[range].value

def mk_first_dir(filename,data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    zdzl=read_zdzl(filename)
    dir_1=Path(data_dir,base_name)    
    mk_dir(dir_1)
    dir_2=Path(data_dir,base_name,zdzl)
    mk_dir(dir_2) 
    for i in range(1,4):
        dir_3=Path(dir_2,"问题"+str(i))
        mk_dir(dir_3)
    return dir_1

def copy_range(worksheet,source_range="B11:B25",target_range="C11:C25"):
    s=[]
    for cell in worksheet[source_range]:
        v=list(cell)[0].value
        s.append(v)
        
    idx=0
    for cell in worksheet[target_range]:        
        cell[0].value=s[idx]
        idx+=1
    return worksheet

def write_end(filename,name_str=name_str,name_range="B29",date_str=date_str,date_range="C29",data_dir='.\data'):
    base_name=Path(filename).name.split(".")[0]
    write_filename=Path(data_dir,base_name,Path(filename).name)
    if write_filename.exists():
        print("文件：%s-->已经存在，要重写请先删除"%write_filename)
        return
    wb=load_workbook(filename)
    ws_name=wb.sheetnames[0]
    ws = wb[ws_name]
    ws[name_range]=name_str
    ws[date_range]=date_str
    ws=copy_range(ws)

    
    print("写入：",write_filename,name_str,date_str)
    if Path(write_filename).exists():
        Path(write_filename).unlink()
    wb.save(write_filename)

def copy_curfile2_datadir(data_dir='./data/'):
    print("将当前目录下的.xlsx文件copy到指定文件夹%s中"%data_dir)
    file_list=[f for f in Path('.').iterdir() if Path(f).is_file()]
    for file in tqdm(file_list,colour='pink'):
        if (file.name.endswith('.xlsx') and not Path(file).exists()):
            src=file
            dest=Path(data_dir,file.name)
            dest.write_bytes(src.read_bytes()) #for binary files

def mk_dir(path):
    if not Path(path).exists():
        print("建立%s..."%path)
        Path.mkdir(Path(path))
        print(path,"OK!") 
    else:
        print("目录:%s-->已经存在。"%path)

def read_table(filename):
    d={}
    wb=load_workbook(filename)
    ws=wb['Sheet1']    
    d['宗地代码（不动产单元号）']=ws['C11'].value
    zdzl=str(ws['C16'].value)
    county,towns,village,LandParcelNumber =parse_zdzl(zdzl)
    d['乡（镇）']=towns
    d['村']=village
    d['宗地号']=LandParcelNumber
    d['存在问题']=ws['B26'].value
    d['现场核实情况']=ws['C26'].value
    d['处理意见']=ws['B27'].value
    return d
    
def parse_zdzl(s):
    '''解析综地名称字段'''
    county=''
    county_num=0
    if '县'  in s:
        county_num=s.index('县')+1
        county=s[:county_num]    
    
        
    if '乡' in s:
        towns=s[county_num:s.index('乡')]
        village=s[s.index('乡')+1:s.rfind('村')]
    elif  '镇' in s:
        towns=s[county_num:s.index('镇')]
        village=s[s.index('镇')+1:s.rfind('村')]
    else:
        print(s)
        return None
   
    LandParcelNumber=re.findall('[\d]+',s)[0]
    return (county,towns,village,LandParcelNumber)

def process_1(data_dir='./data/'):    
    mk_dir(data_dir) 
    copy_curfile2_datadir()
    file_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_file() and f.name.endswith('.xlsx') and len(f.name)==24)]
    print("共找到%s个数据文件"%len(file_list))
    for file in tqdm(file_list,colour='red',):
        print("读取文件",file)  
        mk_first_dir(Path(file))          
        write_end(file)

def process_2(data_dir='./data/'):
    result=None
    dir_list=[f for f in Path(data_dir).iterdir() if (Path(f).is_dir() and  len(f.name)==19)]  
    pbar = tqdm(dir_list)
    for dir in pbar:
        pbar.set_description("\n%s"%(dir.name))
        print("%s"%dir)
        file=Path(dir,dir.name+".xlsx")
        series=pd.Series(read_table(file))
        print(series)
        if result is None:
            result = pd.DataFrame(series)
        else:
            result=pd.concat([result,series],ignore_index=True,axis=1)
    result=result.T
    result=result.set_index(result.columns[0])
    savefiles_count=len(sorted(Path('.').glob("汇总表*.xlsx")))
    savefile_name="汇总表%s.xlsx"%savefiles_count
    result.to_excel(savefile_name)
    print(result)


def get_1_4_image(img_file):   
    img=Image.open(img_file)
    w,h=img.size
    r_w=int(w*2/3)
    r_h=int(h*2/3)
    box=(0,r_h,r_w,h)
    im_crop= img.crop(box)
    im_crop=red20_other2b(im_crop)
    dest='temp.png'
    res_img = Image.fromarray(im_crop)
    res_img.save(dest)    
    #res_img.show()
    return res_img

def red20_other2b(img):	
    ''' 通过遍历颜色替换程序,只留红色，其余为白
	@param	img:	图像矩阵
	@return				替换后的图像矩阵
	'''
    img_arr = np.asarray(img, dtype=np.double)    
    for i in range(img_arr.shape[1]):
        for j in range(img_arr.shape[0]):
            r,g,b=img_arr[j][i]
            #将白色改成红色
            if r+g+b>750:
               r,g,b=(255,0,0)

            if r>(g+b)*1.8 :
               img_arr[j][i] = (255,0,0)
            else:
                img_arr[j][i]=(0,0,0)
    return np.asarray(img_arr, dtype=np.uint8)
			
def remove_more_then_one(dir=None):
    dir_list=[]
    f_list=dir
    for f in f_list:
        f_parent=str(f.parent)
        if not (f_parent in dir_list):
            dir_list.append(f_parent)
        else:
            f_list.remove(f)
    return f_list    




def get_description(img=None,list_dir='.\images'):
    if img is None:
        dir_list=[f for f in Path(list_dir).glob("**/*") if (f.is_file() and (f.suffix in ['.jpg','.png']))] 
        dir_list=remove_more_then_one(dir_list)
        result=[]
        for file in tqdm(dir_list):
            src=file
            #dest=Path('.','temp'+file.suffix)
            #dest.write_bytes(src.read_bytes()) #for binary files
            get_1_4_image(file)
            print("识别%s"%str(file))
            try:
                #res=reader.readtext(str(dest),detail=0)
                res = reader.readtext('temp.png')
            except :
                print("ERROR")
                continue
            res=get_jwd(res)
            print(res)
            res.append(str(file))
            if len(res)==4:
                result.append(res)
    else:
        im =get_1_4_image(img)
        res_list = reader.readtext("temp.png")
        '''
        res=get_jwd(res_list)
        res.append(img)
        print(res)
        result=[res]
        '''
        res_dict=get_jwd_dict(res_list)
        print(res_dict)
        res_dict['源图片']=img        
        if '海拔' in list(res_dict.keys()):
            res_dict.pop('海拔')
        #result=[v for v in res_dict.values() ]
        result= [res_dict]

    return result

def arr2dict(r):
    tmp={}
    tmp['宗地代码（不动产单元号）']=str(Path(r[-1]).parent).split('\\')[1]
    tmp['经度']=r[0]
    tmp['纬度']=r[1]
    tmp['备注']=r[2]
    if not "问题" in tmp['备注']:
        tmp['备注']=tmp['备注']+"-问题1"


    return tmp


def result2df(res):    
    df=pd.DataFrame({},columns=['宗地代码（不动产单元号）','经度','纬度','备注'])
    tmp={}
    for r in res:
       tmp=arr2dict(r)
       df1=pd.DataFrame(pd.Series(tmp)).T
       df=pd.concat([df,df1],axis=0)
    #df.index=['宗地代码（不动产单元号）']
    return df

def hzb_add_by_index(df_add):
    excel_file="汇总_add_经纬度.xlsx"
    savefiles_count=len(sorted(Path('.').glob("汇总表*.xlsx")))
    if savefiles_count==0:
        process_2()
    savefile_name="汇总表%s.xlsx"%(savefiles_count-1)
    hzb=pd.read_excel(savefile_name)
    #hzb.set_index(['宗地代码（不动产单元号）'],inplace=True) 
    hzb=hzb.drop_duplicates(['宗地代码（不动产单元号）'])   
    df=pd.merge(hzb,df_add,on=['宗地代码（不动产单元号）'],how='left') # left , right , outer
    df.to_excel(excel_file)
    add_columns_by_question(excel_file)
    return df

def add_columns_by_question(excel_file):
    df=pd.read_excel(excel_file)
    df1=df[df['备注'].str.contains("问题1",na=False)]
    df1.set_index('宗地代码（不动产单元号）',inplace=True)
    df2=df[df['备注'].str.contains("问题2",na=False)]
    df2=df2[['宗地代码（不动产单元号）','经度','纬度','备注']]    
    df2.columns=['宗地代码（不动产单元号）','经度2','纬度2','问题2']
    df2.set_index('宗地代码（不动产单元号）',inplace=True)
    df3=df[df['备注'].str.contains("问题3",na=False)]
    df3=df3[['宗地代码（不动产单元号）','经度','纬度','备注']]    
    df3.columns=['宗地代码（不动产单元号）','经度3','纬度3','问题3']
    df3.set_index('宗地代码（不动产单元号）',inplace=True)
    res=pd.merge(df1,df2,on=['宗地代码（不动产单元号）'],how='left')
    res=pd.merge(res,df3,on=['宗地代码（不动产单元号）'],how='left')
    res.to_excel("汇总_add_经纬度2.xlsx")
    return res

def res_add_line(res):
    #print(res)
    y_=0
    res_line=[]
    for cell in res:
        y=cell[0][0][1]
        font_hight=(y-y_)*0.4
        s=cell[1].replace('正在获取中','00.0000')
        s=s.replace('海拔','')
        try:
            s=re.sub('[\"/经度纬：,;:备注虞~^筇臣与_赓`轻瘘笫。孥W%@庹 畚-]','',s)
            #print(cell[1],s)
        except:
            print(cell[1])
            continue
        s=s.replace('=','')
        if len(s)==0 or ("米" in s):
            continue
        if (y>(y_+font_hight)) or (y<(y_ - font_hight)):
            res_line.append(s)
        else:
            try:
                res_line.append(res_line[-1]+s)
                res_line.pop(-2)
            except IndexError:
                print(res_line)
                #print(res)
        y_=y
        print(res_line[-1])
    while len(res_line[-1])<8:
        last_line=res_line.pop()    
        last_line=res_line.pop()+last_line
        res_line.append(last_line)
    
    last_line=res_line.pop()    
    last_line=last_line.replace('=','')
    last_line=last_line.replace("'",'')
    last_line=last_line.replace('.','')
    last_line=last_line.replace('问魉','问题')
    last_line=last_line.replace('向题','问题')
    last_line=last_line.replace('问瘿','问题')
    last_line=last_line.replace('问题','-问题')
    last_line=last_line.replace('长校锐','长校镇')
    last_line=last_line.replace('江场','江坊')
    res_line.append(last_line)
    return res_line



def get_jwd(res):  
    res=res_add_line(res)
    if not "问题" in res[-1]:
        res[-1]=res[-1]+"-问题1"
    return res

def get_jwd_dict(res):  
   
    res= res_add_line(res) 
    result_dict={}       
    if len(res)==3 or  len(res)==4:
        try:
            result_dict['经度']=re.findall('[\d.]+',res[0])[0]
            result_dict['纬度']=re.findall('[\d.]+',res[1])[0]
      
            if len(res)==4:
                result_dict['海拔']=re.findall('[\d.]+',res[2])[0]
            result_dict['备注']=res[-1]
        except:
            print("\n\n跳过识别不了的",res,'\n\n')
    return result_dict

def open_last_excelfile(basename):
    savefiles_count=len(sorted(Path('.').glob("%s*.xlsx"%basename)))
    if savefiles_count==0:
        return None
    savefile_name="%s%s.xlsx"%(basename,savefiles_count-1)
    df=pd.read_excel(savefile_name)
    print(df)
    return df

def put_images(img_dir='images'):
    df_hzb=open_last_excelfile(basename='汇总表')
    if df_hzb is None:
        process_2()

    if not Path(img_dir).exists():
        Path(img_dir).mkdir()

    dir_list=[f for f in Path(img_dir).glob("**/*") if (f.is_file() and (f.suffix in ['.jpg','.png']))] 
    if len(dir_list)==0:    
        print("请将要投放的照片文件放入当前目录下的%s目录"%img_dir)
        exit()
    else:
        print("共有%s张图片要处理"%len(dir_list))
    
    result=[]
    for img in tqdm(dir_list):
        res=get_description(img=img)[0]
        if not  '备注' in list(res.keys()):
            continue
        try:
            s=res['备注']
            if not "问题" in s:
                res['备注']=res['备注']+"-问题1"

            area,town,village,landParcelNumber =parse_zdzl(s=s)                        
            index_id=from_dfhzb_get_index(df=df_hzb,town=town,village=village,landParcelNumber=landParcelNumber)
            #res.append(index_id) 
            res['宗地代码（不动产单元号）']=  index_id
            result.append(res)
        except:
            print("\n\n\n\n跳过识别出错的%s\n\n\n"%img)
        else:
            move_pic_form_series(res)  
   
    df=pd.DataFrame(result,columns=['经度','纬度','备注','源图片','宗地代码（不动产单元号）'])
    df.set_index('宗地代码（不动产单元号）')
    print(df)
    
    try:
        df.to_excel(str(Path(img_dir,'list.xlsx')) )
    except Exception as e:
        print(e)

    df_add=df.filter(['经度','纬度','备注','宗地代码（不动产单元号）'])
    df_add.set_index('宗地代码（不动产单元号）')
    hzb_add_by_index(df_add)


    
    

def from_dfhzb_get_index(df=None,town=None,village=None,landParcelNumber=None):
    '''从df文件中获取 索引号，根据 乡，村，宗地号'''
    if df is None:
        df=pd.read_excel('汇总表0.xlsx')
    df=df[df['乡（镇）']==town]    
    df=df[df['村']==village]
    print(town,village,landParcelNumber)
    df=df[df['宗地号']==int(landParcelNumber)]
    #print(landParcelNumber,df)
    print(df['宗地代码（不动产单元号）'])
    return (df.values[0][0])

def move_pic_form_df(df=None):
    for i in range(len(df)):
        serie =df.iloc[i]
        move_pic_form_series(series=serie)

def move_pic_form_series(series=None):        
        src=series['源图片']
        cd1=series['宗地代码（不动产单元号）']
        if not  Path('data',cd1).exists():
            Path('data',cd1).mkdir()
        cd2=series['备注'].split('-')[0]
        if '清流县' not in cd2:
            cd2='清流县'+cd2
        if not  Path('data',cd1,cd2).exists():
            Path('data',cd1,cd2).mkdir()
        try:
            cd3=series['备注'].split('-')[1]
        except:
            cd3=series['备注'][-3:]
        if not  Path('data',cd1,cd2,cd3).exists():
            Path('data',cd1,cd2,cd3).mkdir()
        fn=Path(src).name
        dest =Path('data',cd1,cd2,cd3,fn)
        try:
            shutil.move(src,dest)
        except:
            print("移动文件%s失败"%src)
        else:
            print("移动文件OK!%s->\n%s"%(src,dest))



if __name__=="__main__": 
    if len(sys.argv)==2:
        num=str(sys.argv[1])
    else:
        prog_name=Path(sys.argv[0]).name 
        num=prog_name[0]

    if num=='1': 
       process_1() 
    elif num=='2': 
       process_2()
    elif num=='3': 
        
        import easyocr 
        reader = easyocr.Reader(['ch_sim','en'],recog_network ='zh_sim_g2') # this needs to run only once to load the model into memory
        result =get_description(list_dir='.\data')
        res_df =result2df(result)
        print(res_df)
        hzb_add_by_index(res_df)
    elif num=="4":
        import easyocr 
        reader = easyocr.Reader(['ch_sim','en'],recog_network ='zh_sim_g2')
        put_images()


        
    else:
        print("Usage: \n%s"%doc)